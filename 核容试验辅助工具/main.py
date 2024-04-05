import sys
import os
from pathlib import Path
from PySide6.QtWidgets import QApplication, QFileDialog, QWidget, QMessageBox
from run_widget import Ui_Form
from openpyxl import load_workbook
import random


class LoginWindow(Ui_Form, QWidget):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.initUI()

    def initUI(self):
        self.FILE220_remain = 'views/1浪头220kV第一组蓄电池充放电记录.xlsx'  # 220自己留存的模版
        self.FILE66_submit = 'views/五龙背变蓄电池记录.xlsx'  # 66提交监理的模版
        self.FILE66_remain = 'views/团山66kV蓄电池充放电记录.xlsx'  # 66自己留存的模版
        self.filePath = None
        self.filepath_choose_tag = False  # 如果自选模版，置为True
        self.current_radioButton = 'radioButton_remain'  # 当前是哪个单选框
        self.voltage = '220kV'
        self.workbook = None
        self.name = '天华'
        self.person = None
        self.choice = '第一组'
        self.date = None
        self.date_dict = {'year': None, 'month': None, 'day': None}
        self.time = None
        self.time_dict = {'hour': None, 'minute': None}
        self.sheet_name_220_remain = [
            '封页',
            '放电前',
            '放电1小时',
            '放电2小时',
            '放电3小时',
            '放电4小时',
            '放电5小时',
            '放电6小时',
            '放电7小时',
            '放电8小时',
            '放电9小时',
            '放电10小时',
            '充电前',
            '充电1小时',
            '充电2小时',
            '充电3小时',
            '充电4小时',
            '充电5小时',
            '充电6小时',
            '充电7小时',
            '充电8小时',
            '充电9小时',
            '充电10小时']
        self.sheet_name_66_remain = [
            '封页',
            '放电前',
            '放电1小时',
            '放电2小时',
            '放电3小时',
            '放电4小时',
            '放电5小时',
            '充电前',
            '充电1小时',
            '充电2小时',
            '充电3小时',
            '充电4小时',
            '充电5小时']
        self.sheet_name_66_submit = [
            '封皮',
            '放电1小时',
            '放电2小时',
            '放电3小时',
            '放电4小时',
            '放电5小时',
            '充电1小时',
            '充电2小时',
            '充电3小时',
            '充电4小时',
            '充电5小时']
        self.hours = [i for i in range(24)]
        self.battery_capacity = 300
        self.random_value = 0.005
        self.get_error = False

        self.bind()

    def bind(self):
        self.btn_openfile.clicked.connect(self.openFile)
        self.lineEdit_battery_capacity.textChanged.connect(
            lambda: self.save_data('lineEdit_battery_capacity'))  # 填入蓄电池容量
        self.lineEdit_random_value.textChanged.connect(lambda: self.save_data('lineEdit_random_value'))  # 填入变化幅度
        self.comboBox_choose.currentTextChanged.connect(lambda: self.save_data('comboBox_choose'))  # 选择第几组蓄电池
        self.lineEdit_name.textChanged.connect(lambda: self.save_data('lineEdit_name'))  # 填入哪个变电站
        self.dateEdit.dateChanged.connect(lambda: self.save_data('dateEdit'))  # 填入日期
        self.timeEdit.timeChanged.connect(lambda: self.save_data('timeEdit'))  # 填入时间
        self.comboBox_voltage.currentTextChanged.connect(lambda: self.save_data('comboBox_voltage'))  # 选择电压等级
        self.radioButton_submit.clicked.connect(lambda: self.save_data('radioButton_submit'))  # 监理模版点这个
        self.radioButton_remain.clicked.connect(lambda: self.save_data('radioButton_remain'))  # 自己留存点这个
        self.lineEdit_person.textChanged.connect(lambda: self.save_data('lineEdit_person'))  # 自己留存点这个
        self.btn_run.clicked.connect(self.run)

    def fetch_resource(self, rsrc_path):
        # 尝试获取当前程序的资源路径
        try:
            base_path = Path(sys._MEIPASS)  # 获取资源路径
        except AttributeError:
            # 如果不是作为可执行文件运行，则直接返回原始路径
            return rsrc_path  # 没有以可执行文件形式运行，只需返回未修改的路径
        else:
            # 创建一个消息框来显示资源路径
            msg_box = QMessageBox()
            msg_box.setText(str(base_path.joinpath(rsrc_path)))  # 设置消息框文本为资源路径
            msg_box.exec()  # 显示消息框
            # 返回修改后的资源路径
            return base_path.joinpath(rsrc_path)

    def openFile(self):
        # 创建一个QFileDialog
        # options = QFileDialog.Options()
        try:

            self.filePath, _ = QFileDialog.getOpenFileName(self, "打开文件", "",
                                                           "所有文件 (*);；文本文件 (*.xls)；文本文件 (*.xlsx)")

            if self.filePath:
                self.filepath_choose_tag = True
            # self.workbook = load_workbook(self.filePath)


        except:
            self.show_message('请再次检查文件路径')

    def save_data(self, tag):
        if tag == 'lineEdit_name':
            self.name = self.lineEdit_name.text()
        elif tag == 'lineEdit_battery_capacity':
            try:
                self.battery_capacity = int(self.lineEdit_battery_capacity.text())
            except ValueError:
                self.show_message('蓄电池组容量输入错误')
        elif tag == 'lineEdit_random_value':
            try:
                self.random_value = float(self.lineEdit_random_value.text())
            except ValueError:
                self.show_message('变化幅值输入错误')
        elif tag == 'comboBox_choose':
            self.choice = self.comboBox_choose.currentText()
        elif tag == 'dateEdit':
            date = self.dateEdit.date()
            year = date.year()
            month = date.month()
            day = date.day()
            self.date_dict = {'year': year, 'month': month, 'day': day}
        elif tag == 'lineEdit_person':
            self.person = self.lineEdit_person.text()
        elif tag == 'timeEdit':
            time = self.timeEdit.time()
            hour = time.hour()
            minute = time.minute()
            self.time_dict = {'hour': hour, 'minute': minute}
        elif tag == 'comboBox_voltage':
            self.voltage = self.comboBox_voltage.currentText()
        elif tag == 'radioButton_submit':
            self.current_radioButton = tag
        elif tag == 'radioButton_remain':
            self.current_radioButton = tag

        print(self.name, self.choice, self.date_dict, self.time, self.current_radioButton, self.voltage)

    def run(self):

        try:
            # 先判断是不是已经在openfile里选好了路径：
            if not self.filepath_choose_tag:
                if self.current_radioButton == 'radioButton_remain':
                    if self.voltage == '220kV':
                        self.filePath = self.fetch_resource(self.FILE220_remain)
                        print('1')
                    elif self.voltage == '66kV':
                        self.filePath = self.fetch_resource(self.FILE66_remain)
                        print('2')
                elif self.current_radioButton == 'radioButton_submit':
                    self.filePath = self.fetch_resource(self.FILE66_submit)
                    print('3')
            # elif self.battery_capacity == None:
            #     self.show_message('请输入蓄电池容量')

            elif self.random_value == None:
                self.show_message('请输入变化幅度')
            elif self.random_value > 0.1:
                self.show_message('变化幅度太大')
            elif self.name == '':
                self.show_message('请输入名称')
            elif self.choice not in ['第一组', '第二组']:
                self.show_message('请选择蓄电池组')
            elif self.date_dict['year'] is None:
                self.show_message('请选择日期')
            elif self.time_dict['hour'] is None:
                self.show_message('请输入时间')
            print(self.filePath)
            self.workbook = load_workbook(self.filePath)
            if self.current_radioButton == 'radioButton_remain' and self.voltage == '220kV':
                self.first_page_220_remain()
                self.xlsx_change_220_remain1()
                self.xlsx_change_220_remain2()
                self.filename = f'1{self.name}220kV{self.choice}蓄电池充放电记录.xlsx'
            elif self.current_radioButton == 'radioButton_remain' and self.voltage == '66kV':
                self.first_page_66_remain()
                self.xlsx_change_66_remain1()
                self.xlsx_change_66_remain2()
                self.filename = f'{self.name}变66kV蓄电池充放电记录.xlsx'
            elif self.current_radioButton == 'radioButton_submit' and self.voltage == '66kV':
                self.first_page_66_submit()
                self.xlsx_change_66_submit1()
                # self.xlsx_change_66_submit2()
                self.filename = f'{self.name}变66kV蓄电池充放电记录.xlsx'
            # 保存修改后的工作簿
            desktop_path = os.path.join(
                sys.platform == "win32" and os.path.join(os.environ['USERPROFILE'], 'Desktop') or os.path.join(
                    os.environ['HOME'], 'Desktop'))
            print(desktop_path)

            # 保存修改后的工作簿到桌面

            self.workbook.save(os.path.join(desktop_path, self.filename))
            self.workbook.close()
            # self.workbook.save('11111111.xlsx')
        except Exception as e:
            self.show_message(f'发生了一个错误： {e}')
            self.show_message('请再次检查参数')

    def show_message(self, info):
        # self.get_error = True
        msgBox = QMessageBox()
        # 设置消息框的标题
        msgBox.setWindowTitle('提示')
        # 设置消息框的文本内容
        msgBox.setText(info)
        msgBox.exec()

    def first_page_220_remain(self):
        worksheet = self.workbook.worksheets[0]

        '''变电站名'''
        cell_data_a8 = 'a8'
        new_value_a8 = f'{self.name}220kV变电站'
        worksheet[cell_data_a8] = new_value_a8
        '''蓄电池组'''
        cell_data_d31 = 'd31'
        new_value_d31 = '一套蓄电池' if self.choice == '第一组' else '二套蓄电池'
        worksheet[cell_data_d31] = new_value_d31
        '''年份'''
        cell_data_d32 = 'd32'
        new_value_d32 = f'{self.date_dict["year"]}年 '
        worksheet[cell_data_d32] = new_value_d32

    def xlsx_change_220_remain1(self):

        for i in range(1, 23):
            # 使用openpyxl加载工作簿
            '''日期：'''
            worksheet = self.workbook.worksheets[i]
            cell_data_a10 = 'A10'
            # 判断日期：
            new_value_a10 = self.date_dict['day'] + 1 if i > 10 and worksheet['B10'].value <= 11 else self.date_dict[
                'day']
            # 设置新的单元格值
            worksheet[cell_data_a10] = new_value_a10
            '''日期：'''
            cell_data_I2 = 'I2'  # 室温：25℃      2024    年 3   月  25   日
            new_value_I2 = f'室温：25℃      {self.date_dict["year"]}    年 {self.date_dict["month"]}   月  {new_value_a10}   日'  # 室温：25℃      2024    年 3   月  26   日
            # 设置新的单元格值
            worksheet[cell_data_I2] = new_value_I2
            '''变电站名'''
            cell_data_a2 = 'a2'  # 室温：25℃      2024    年 3   月  25   日
            new_value_a2 = f'蓄电池名称：{self.name}220kV变电站{self.choice}蓄电池'
            worksheet[cell_data_a2] = new_value_a2

            '''蓄电池容量'''
            cell_data_d47 = 'd47'  # 室温：25℃      2024    年 3   月  25   日
            new_value_d47 = f'共   {self.battery_capacity}  安时'
            worksheet[cell_data_d47] = new_value_d47

            '''人名'''

            cell_data_K47 = 'K47'  # 室温：25℃      2024    年 3   月  25   日
            new_value_K47 = self.person
            worksheet[cell_data_K47] = new_value_K47

            '''时间'''
            cell_data_b10 = 'b10'
            index = self.time_dict['hour'] + i - 1 if self.time_dict['hour'] + i - 1 < 24 else self.time_dict[
                                                                                                   'hour'] + i - 1 - 24
            new_value_b10 = self.hours[index]
            worksheet[cell_data_b10] = new_value_b10
            cell_data_c10 = 'c10'
            new_value_c10 = self.time_dict['minute']
            worksheet[cell_data_c10] = new_value_c10

    def xlsx_change_220_remain2(self):

        for i in range(1, 23):
            # 使用openpyxl加载工作簿
            total_sum = 0
            '''电压：'''
            worksheet = self.workbook.worksheets[i]
            # 遍历H5到H46的单元格
            for row in range(5, 47):  # 从第5行开始，到第46行结束
                cell_value = worksheet.cell(row=row, column=8).value  # H列是第8列
                if cell_value is not None:  # 确保单元格有值
                    try:
                        # 尝试将单元格的值转换为浮点数
                        cell_value = float(cell_value)
                        # 生成负数随机数并相加
                        corrected_value = round(cell_value + random.uniform(self.random_value, -self.random_value), 3)
                        # 将改正后的值写入对应的单元格
                        worksheet.cell(row=row, column=8, value=corrected_value)
                        total_sum += corrected_value
                    except ValueError:
                        # 如果转换失败，说明单元格中的值不是数字，不做任何改动
                        worksheet.cell(row=row, column=8, value=cell_value)

                cell_value = worksheet.cell(row=row, column=11).value  # K列是第11列
                if cell_value is not None:  # 确保单元格有值
                    try:
                        # 尝试将单元格的值转换为浮点数
                        cell_value = float(cell_value)
                        # 生成负数随机数并相加
                        corrected_value = round(cell_value + random.uniform(self.random_value, -self.random_value), 3)
                        # 将改正后的值写入对应的单元格
                        worksheet.cell(row=row, column=11, value=corrected_value)
                        total_sum += corrected_value
                    except ValueError:
                        # 如果转换失败，说明单元格中的值不是数字，不做任何改动
                        self.show_message(f"Cell value {i}is {row}not a number: {cell_value}")
                        worksheet.cell(row=row, column=11, value=cell_value)

                cell_value = worksheet.cell(row=row, column=14).value  # N列是第14列
                if cell_value is not None:  # 确保单元格有值
                    try:
                        # 尝试将单元格的值转换为浮点数
                        cell_value = float(cell_value)
                        # 生成负数随机数并相加
                        corrected_value = round(cell_value + random.uniform(self.random_value, -self.random_value), 3)
                        # 将改正后的值写入对应的单元格
                        worksheet.cell(row=row, column=14, value=corrected_value)
                        total_sum += corrected_value
                    except ValueError:
                        # 如果转换失败，说明单元格中的值不是数字，不做任何改动
                        self.show_message(f"Cell value {i}is {row}not a number: {cell_value}")
                        worksheet.cell(row=row, column=14, value=cell_value)

            cell_data_d10 = 'd10'
            new_value_d10 = total_sum
            worksheet[cell_data_d10] = new_value_d10

    def first_page_66_remain(self):

        worksheet = self.workbook.worksheets[0]

        '''变电站名'''
        cell_data_a8 = 'a8'
        new_value_a8 = f'{self.name}66kV变电站'
        worksheet[cell_data_a8] = new_value_a8

        '''年份'''
        cell_data_d32 = 'd32'
        new_value_d32 = f'{self.date_dict["year"]}年 '
        worksheet[cell_data_d32] = new_value_d32

    def xlsx_change_66_remain1(self):

        for i in range(1, 13):
            # 使用openpyxl加载工作簿
            '''日期：'''
            worksheet = self.workbook.worksheets[i]
            cell_data_a10 = 'A10'
            # 判断日期：
            new_value_a10 = self.date_dict['day'] + 1 if i > 10 and worksheet['B10'].value <= 11 else self.date_dict[
                'day']
            # 设置新的单元格值
            worksheet[cell_data_a10] = new_value_a10
            '''日期：'''
            cell_data_I2 = 'I2'  # 室温：25℃      2024    年 3   月  25   日
            new_value_I2 = f'室温：25℃      {self.date_dict["year"]}    年 {self.date_dict["month"]}   月  {new_value_a10}   日'  # 室温：25℃      2024    年 3   月  26   日
            # 设置新的单元格值
            worksheet[cell_data_I2] = new_value_I2
            '''变电站名'''
            cell_data_a2 = 'a2'  # 室温：25℃      2024    年 3   月  25   日
            new_value_a2 = f'蓄电池名称：{self.name}66kV变电站蓄电池'
            worksheet[cell_data_a2] = new_value_a2

            '''蓄电池容量'''
            cell_data_d47 = 'd47'  # 室温：25℃      2024    年 3   月  25   日
            new_value_d47 = f'共   {self.battery_capacity}  安时'
            worksheet[cell_data_d47] = new_value_d47

            '''时间'''
            cell_data_b10 = 'b10'
            index = self.time_dict['hour'] + i - 1 if self.time_dict['hour'] + i - 1 < 24 else self.time_dict[
                                                                                                   'hour'] + i - 1 - 24
            new_value_b10 = self.hours[index]
            worksheet[cell_data_b10] = new_value_b10
            cell_data_c10 = 'c10'
            new_value_c10 = self.time_dict['minute']
            worksheet[cell_data_c10] = new_value_c10

    def xlsx_change_66_remain2(self):

        for i in range(1, 13):
            # range是左闭右开区间
            total_sum = 0
            '''电压：'''
            worksheet = self.workbook.worksheets[i]
            # 遍历H5到H46的单元格
            for row in range(5, 47):  # 从第5行开始，到第46行结束
                cell_value = worksheet.cell(row=row, column=8).value  # H列是第8列
                if cell_value is not None:  # 确保单元格有值
                    try:
                        # 尝试将单元格的值转换为浮点数
                        cell_value = float(cell_value)
                        # 生成负数随机数并相加
                        corrected_value = round(cell_value + random.uniform(self.random_value, -self.random_value), 3)
                        # 将改正后的值写入对应的单元格
                        worksheet.cell(row=row, column=8, value=corrected_value)
                        total_sum += corrected_value
                    except ValueError:
                        # 如果转换失败，说明单元格中的值不是数字，不做任何改动，用模版不存在这种情况
                        worksheet.cell(row=row, column=8, value=cell_value)

                cell_value = worksheet.cell(row=row, column=11).value  # K列是第11列
                if cell_value is not None:  # 确保单元格有值
                    try:
                        # 尝试将单元格的值转换为浮点数
                        cell_value = float(cell_value)
                        # 生成负数随机数并相加
                        corrected_value = round(cell_value + random.uniform(self.random_value, -self.random_value), 3)
                        # 将改正后的值写入对应的单元格
                        worksheet.cell(row=row, column=11, value=corrected_value)
                        total_sum += corrected_value
                    except ValueError:
                        # 如果转换失败，说明单元格中的值不是数字，不做任何改动
                        self.show_message(f"Cell value {i}is {row}not a number: {cell_value}")
                        worksheet.cell(row=row, column=11, value=cell_value)

                cell_value = worksheet.cell(row=row, column=14).value  # N列是第14列
                if cell_value is not None:  # 确保单元格有值
                    try:
                        # 尝试将单元格的值转换为浮点数
                        cell_value = float(cell_value)
                        # 生成负数随机数并相加
                        corrected_value = round(cell_value + random.uniform(self.random_value, -self.random_value), 3)
                        # 将改正后的值写入对应的单元格
                        worksheet.cell(row=row, column=14, value=corrected_value)
                        total_sum += corrected_value
                    except ValueError:
                        # 如果转换失败，说明单元格中的值不是数字，不做任何改动
                        self.show_message(f"Cell value {i}is {row}not a number: {cell_value}")
                        worksheet.cell(row=row, column=14, value=cell_value)

            cell_data_d10 = 'd10'
            new_value_d10 = total_sum
            worksheet[cell_data_d10] = new_value_d10

    def first_page_66_submit(self):

        worksheet = self.workbook.worksheets[0]

        '''变电站名'''
        cell_data_a8 = 'a8'
        new_value_a8 = f'{self.name}66kV变电站'
        worksheet[cell_data_a8] = new_value_a8

        '''年份'''
        cell_data_d32 = 'd32'
        new_value_d32 = f'{self.date_dict["year"]}年 '
        worksheet[cell_data_d32] = new_value_d32

    def xlsx_change_66_submit1(self):

        for i in range(1, 11):
            worksheet = self.workbook.worksheets[i]
            '''充电电流'''
            I10 = int(self.battery_capacity / 10)
            '''时间'''
            index = self.time_dict['hour'] + i - 1
            if i >= 6:
                index += 1
            total_sum = 0
            total_min = 2.3
            '''电压：'''
            for row in range(4, 46):  # 从第5行开始，到第46行结束
                cell_value = worksheet.cell(row=row, column=2).value  # H列是第8列
                if cell_value is not None:  # 确保单元格有值
                    try:
                        # 尝试将单元格的值转换为浮点数
                        cell_value = float(cell_value)
                        # 生成负数随机数并相加
                        corrected_value = round(cell_value + random.uniform(self.random_value, -self.random_value), 3)
                        # 将改正后的值写入对应的单元格
                        worksheet.cell(row=row, column=2, value=corrected_value)
                        if total_min > corrected_value:
                            total_min = corrected_value
                        total_sum += corrected_value
                    except ValueError:
                        # 如果转换失败，说明单元格中的值不是数字，不做任何改动，用模版不存在这种情况
                        worksheet.cell(row=row, column=2, value=cell_value)

                cell_value = worksheet.cell(row=row, column=5).value  # K列是第11列
                if cell_value is not None:  # 确保单元格有值
                    try:
                        # 尝试将单元格的值转换为浮点数
                        cell_value = float(cell_value)
                        # 生成负数随机数并相加
                        corrected_value = round(cell_value + random.uniform(self.random_value, -self.random_value), 3)
                        # 将改正后的值写入对应的单元格
                        worksheet.cell(row=row, column=5, value=corrected_value)
                        if total_min > corrected_value:
                            total_min = corrected_value
                        total_sum += corrected_value
                    except ValueError:
                        # 如果转换失败，说明单元格中的值不是数字，不做任何改动
                        self.show_message(f"Cell value {i}is {row}not a number: {cell_value}")
                        worksheet.cell(row=row, column=5, value=cell_value)

                cell_value = worksheet.cell(row=row, column=8).value  # N列是第14列
                if cell_value is not None:  # 确保单元格有值
                    try:
                        # 尝试将单元格的值转换为浮点数
                        cell_value = float(cell_value)
                        # 生成负数随机数并相加
                        corrected_value = round(cell_value + random.uniform(self.random_value, -self.random_value), 3)
                        # 将改正后的值写入对应的单元格
                        worksheet.cell(row=row, column=8, value=corrected_value)
                        if total_min > corrected_value:
                            total_min = corrected_value
                        total_sum += corrected_value
                    except ValueError:
                        # 如果转换失败，说明单元格中的值不是数字，不做任何改动
                        self.show_message(f"Cell value {i}is {row}not a number: {cell_value}")
                        worksheet.cell(row=row, column=8, value=cell_value)
            total_min = round(total_min, 3)
            total_sum = round(total_sum, 1)
            cell_data_a46 = 'a46'
            new_value_a46 = f'''放电电流：{I10}A    放电电压：{total_sum}V    单体电池最低终止电压：{total_min}V    环境温度：25℃\n测量时间：{self.date_dict["year"]}年{self.date_dict["month"]}月{self.date_dict["day"]}日{self.hours[index]}时{self.time_dict['minute']}分  '''
            worksheet[cell_data_a46] = new_value_a46


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = LoginWindow()
    ex.show()
    sys.exit(app.exec_())
